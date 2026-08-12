"""
Runs the analyses described in README.md against the trimmed dataset and produces:
  - outputs/chart1_top_corridors.png
  - outputs/chart2_provider_comparison.png
  - outputs/chart3_cost_trend.png
  - outputs/chart4_corridor_heatmap.png
  - outputs/chart5_cheapest_vs_expensive.png
  - outputs/chart6_choropleth.png + outputs/chart6_choropleth.html (interactive)
  - outputs/chart7_payment_flow_diagram.png (explainer: how the money actually moves)
  - analysis/CrossBorderPayments_Analysis.xlsx (summary tables + native Excel charts
    + conditional-formatting heatmap)

Minimum observation count per corridor/group is enforced throughout so a single
one-off quote can't distort an average.
"""

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import pandas as pd
import plotly.express as px
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font

DATA_PATH = "data/remittance_prices_selected_columns.csv"
OUTPUTS_DIR = "outputs"
WORKBOOK_PATH = "analysis/CrossBorderPayments_Analysis.xlsx"

MIN_OBSERVATIONS = 5
TOP_N = 10
HEATMAP_TOP_COUNTRIES = 15

plt.rcParams.update(
    {
        "figure.facecolor": "white",
        "axes.facecolor": "white",
        "axes.edgecolor": "#5a6c70",
        "axes.labelcolor": "#102428",
        "text.color": "#102428",
        "xtick.color": "#102428",
        "ytick.color": "#102428",
        "font.size": 10,
    }
)
TEAL = "#087c7f"
CORAL = "#ff6f61"
TEAL_DARK = "#05363d"


def load_data() -> pd.DataFrame:
    df = pd.read_csv(DATA_PATH)
    df["corridor_label"] = df["source_name"] + " -> " + df["destination_name"]

    def period_sort_key(period: str) -> tuple[int, int]:
        year, quarter = period.split("_")
        return int(year), int(quarter[0])

    df["period_sort"] = df["period"].map(period_sort_key)
    return df


def corridor_averages(df: pd.DataFrame) -> pd.DataFrame:
    # groupby(...) bundles every row for the same corridor (e.g. all "UK -> India"
    # quotes) together; .agg(["mean", "count"]) then computes the average cost and
    # how many quotes went into it, for every corridor at once - this is the same
    # thing an Excel PivotTable does when you drag a field into "Values" twice.
    grouped = df.groupby("corridor_label")["total_cost_pct"].agg(["mean", "count"])
    return grouped[grouped["count"] >= MIN_OBSERVATIONS].sort_values("mean")


def chart1_top_corridors(df: pd.DataFrame) -> pd.DataFrame:
    corridors = corridor_averages(df)
    expensive = corridors.tail(TOP_N).sort_values("mean")
    cheap = corridors.head(TOP_N).sort_values("mean", ascending=False)

    fig, axes = plt.subplots(1, 2, figsize=(12, 5))
    axes[0].barh(cheap.index, cheap["mean"], color=TEAL)
    axes[0].set_title(f"{TOP_N} cheapest corridors")
    axes[0].set_xlabel("Avg. total cost (%)")

    axes[1].barh(expensive.index, expensive["mean"], color=CORAL)
    axes[1].set_title(f"{TOP_N} most expensive corridors")
    axes[1].set_xlabel("Avg. total cost (%)")

    fig.suptitle("Cost to send $200-equivalent, by corridor")
    fig.tight_layout()
    fig.savefig(f"{OUTPUTS_DIR}/chart1_top_corridors.png", dpi=150)
    plt.close(fig)
    return corridors


def chart2_provider_comparison(df: pd.DataFrame) -> pd.Series:
    by_type = df.groupby("firm_type")["total_cost_pct"].agg(["mean", "count"])
    by_type = by_type[by_type["count"] >= MIN_OBSERVATIONS].sort_values("mean")

    fig, ax = plt.subplots(figsize=(9, 5))
    ax.barh(by_type.index, by_type["mean"], color=TEAL)
    ax.set_xlabel("Avg. total cost (%)")
    ax.set_title("Average cost by provider type")
    fig.tight_layout()
    fig.savefig(f"{OUTPUTS_DIR}/chart2_provider_comparison.png", dpi=150)
    plt.close(fig)
    return by_type["mean"]


def chart3_cost_trend(df: pd.DataFrame) -> pd.DataFrame:
    top_types = df["firm_type"].value_counts().head(3).index
    trend = (
        df[df["firm_type"].isin(top_types)]
        .groupby(["period_sort", "period", "firm_type"])["total_cost_pct"]
        .mean()
        .reset_index()
        .sort_values("period_sort")
    )
    # .pivot(...) reshapes the long list of (period, firm_type, cost) rows into a
    # wide table - one row per quarter, one column per firm type - which is exactly
    # the shape a line chart needs (one line per column).
    pivot = trend.pivot(index="period", columns="firm_type", values="total_cost_pct")
    pivot = pivot.reindex(trend.drop_duplicates("period").sort_values("period_sort")["period"])

    fig, ax = plt.subplots(figsize=(12, 5))
    pivot.plot(ax=ax, marker="o", markersize=3)
    ax.set_ylabel("Avg. total cost (%)")
    ax.set_title("Cost trend over time, by provider type")
    ax.tick_params(axis="x", rotation=90, labelsize=6)
    fig.tight_layout()
    fig.savefig(f"{OUTPUTS_DIR}/chart3_cost_trend.png", dpi=150)
    plt.close(fig)
    return pivot


def chart4_corridor_heatmap(df: pd.DataFrame) -> pd.DataFrame:
    top_sources = df["source_name"].value_counts().head(HEATMAP_TOP_COUNTRIES).index
    top_dests = df["destination_name"].value_counts().head(HEATMAP_TOP_COUNTRIES).index
    subset = df[df["source_name"].isin(top_sources) & df["destination_name"].isin(top_dests)]
    # pivot_table is like pivot() above but also aggregates (here: averages) any
    # duplicate (sending country, receiving country) combinations as it reshapes -
    # the result is a sending-country x receiving-country grid ready to colour in.
    matrix = subset.pivot_table(
        index="source_name", columns="destination_name", values="total_cost_pct", aggfunc="mean"
    )

    fig, ax = plt.subplots(figsize=(11, 9))
    im = ax.imshow(matrix.values, cmap="RdYlGn_r", aspect="auto")
    ax.set_xticks(range(len(matrix.columns)))
    ax.set_xticklabels(matrix.columns, rotation=90, fontsize=7)
    ax.set_yticks(range(len(matrix.index)))
    ax.set_yticklabels(matrix.index, fontsize=7)
    ax.set_title(f"Avg. cost (%) - top {HEATMAP_TOP_COUNTRIES} sending x receiving countries by data volume")
    fig.colorbar(im, ax=ax, label="Avg. total cost (%)")
    fig.tight_layout()
    fig.savefig(f"{OUTPUTS_DIR}/chart4_corridor_heatmap.png", dpi=150)
    plt.close(fig)
    return matrix


def chart5_cheapest_vs_expensive(df: pd.DataFrame, corridors: pd.DataFrame) -> None:
    cheap5 = corridors.head(5).index
    expensive5 = corridors.tail(5).index
    subset = df[df["corridor_label"].isin(list(cheap5) + list(expensive5))]

    top_types = subset["firm_type"].value_counts().head(3).index
    pivot = (
        subset[subset["firm_type"].isin(top_types)]
        .groupby(["corridor_label", "firm_type"])["total_cost_pct"]
        .mean()
        .unstack()
        .reindex(list(cheap5) + list(expensive5))
    )

    fig, ax = plt.subplots(figsize=(12, 6))
    pivot.plot(kind="bar", ax=ax)
    ax.set_ylabel("Avg. total cost (%)")
    ax.set_title("5 cheapest vs 5 most expensive corridors, by provider type")
    ax.tick_params(axis="x", rotation=45, labelsize=8)
    fig.tight_layout()
    fig.savefig(f"{OUTPUTS_DIR}/chart5_cheapest_vs_expensive.png", dpi=150)
    plt.close(fig)


def chart6_choropleth(df: pd.DataFrame) -> pd.DataFrame:
    # This is the cost to RECEIVE money in each country, averaged across every
    # sending country and provider that quotes a price into it - "how expensive is
    # it to be on the receiving end of a remittance here?" grouping by
    # destination_code (not destination_name) because that's what plotly's map
    # needs to know where to draw the shading.
    by_country = df.groupby(["destination_code", "destination_name"])["total_cost_pct"].agg(
        ["mean", "count"]
    ).reset_index()
    by_country = by_country[by_country["count"] >= MIN_OBSERVATIONS]

    fig = px.choropleth(
        by_country,
        locations="destination_code",
        locationmode="ISO-3",
        color="mean",
        hover_name="destination_name",
        hover_data={"destination_code": False, "count": True, "mean": ":.2f"},
        color_continuous_scale="RdYlGn_r",
        labels={"mean": "Avg. total cost (%)", "count": "Observations"},
        title="Average cost to receive a remittance, by country",
    )
    fig.update_layout(
        font_color="#102428",
        title_font_size=18,
        margin=dict(l=10, r=10, t=60, b=10),
        coloraxis_colorbar_title="Avg. cost (%)",
    )
    fig.write_image(f"{OUTPUTS_DIR}/chart6_choropleth.png", width=1400, height=800, scale=2)
    fig.write_html(f"{OUTPUTS_DIR}/chart6_choropleth.html", include_plotlyjs="cdn")
    return by_country


def chart7_payment_flow_diagram(df: pd.DataFrame) -> None:
    """
    A plain-English explainer, not a data chart: shows why a bank transfer and a
    mobile/digital transfer end up costing such different amounts for the same
    $200. Each box the money passes through is a place a fee or FX margin can be
    added - more boxes (the bank path) means more places to lose money.
    """
    bank_avg = df[df["firm_type"] == "Bank"]["total_cost_pct"].mean()
    mobile_avg = df[df["firm_type"] == "Mobile Operator"]["total_cost_pct"].mean()

    fig, ax = plt.subplots(figsize=(12, 6))
    ax.set_xlim(0, 12)
    ax.set_ylim(0, 5.6)
    ax.axis("off")

    def box(x: float, y: float, w: float, label: str, color: str) -> None:
        ax.add_patch(
            plt.Rectangle((x, y), w, 0.9, facecolor=color, edgecolor=TEAL_DARK, linewidth=1.2, zorder=2)
        )
        ax.text(x + w / 2, y + 0.45, label, ha="center", va="center", fontsize=9, color="white", zorder=3, wrap=True)

    def arrow(x_from: float, x_to: float, y: float, note: str = "") -> None:
        ax.annotate(
            "", xy=(x_to, y), xytext=(x_from, y),
            arrowprops=dict(arrowstyle="-|>", color=TEAL_DARK, linewidth=1.4),
            zorder=1,
        )
        if note:
            ax.text((x_from + x_to) / 2, y + 0.18, note, ha="center", va="bottom", fontsize=7.5, color="#5a6c70")

    # Bank path: more stops between sender and recipient
    y_bank = 3.6
    ax.text(0.1, y_bank + 1.5, "Typical bank transfer", fontsize=11, fontweight="bold", color=TEAL_DARK)
    boxes_bank = [
        (0.1, "Sender"), (2.0, "Sending\nbank"), (3.9, "Correspondent\nbank"),
        (5.8, "Receiving\nbank"), (7.7, "Recipient"),
    ]
    for x, label in boxes_bank:
        box(x, y_bank, 1.7, label, CORAL)
    for (x1, _), (x2, _) in zip(boxes_bank, boxes_bank[1:]):
        arrow(x1 + 1.7, x2, y_bank + 1.0, "fee / FX margin")
    ax.text(9.7, y_bank + 0.45, f"~{bank_avg:.1f}% avg", fontsize=10, fontweight="bold", color=CORAL, va="center")

    # Digital path: sender and recipient's providers already hold local balances,
    # so money moves without a correspondent-bank hop in the middle.
    y_digital = 0.9
    ax.text(0.1, y_digital + 1.5, "Typical mobile / digital transfer", fontsize=11, fontweight="bold", color=TEAL_DARK)
    boxes_digital = [(0.1, "Sender"), (2.0, "Digital\nprovider"), (3.9, "Recipient")]
    for x, label in boxes_digital:
        box(x, y_digital, 1.7, label, TEAL)
    for (x1, _), (x2, _) in zip(boxes_digital, boxes_digital[1:]):
        arrow(x1 + 1.7, x2, y_digital + 1.0, "fee / FX margin")
    ax.text(9.7, y_digital + 0.45, f"~{mobile_avg:.1f}% avg", fontsize=10, fontweight="bold", color=TEAL, va="center")

    fig.suptitle("Same $200 transfer, two different paths", fontsize=13)
    fig.tight_layout()
    fig.savefig(f"{OUTPUTS_DIR}/chart7_payment_flow_diagram.png", dpi=150)
    plt.close(fig)


def build_workbook(
    corridors: pd.DataFrame,
    provider_avg: pd.Series,
    trend: pd.DataFrame,
    heatmap: pd.DataFrame,
) -> None:
    wb = Workbook()
    header_font = Font(bold=True)

    ws1 = wb.active
    ws1.title = "Top-Bottom Corridors"
    ws1.append(["Corridor", "Avg total cost %", "Observations"])
    for cell in ws1[1]:
        cell.font = header_font
    for label, row in pd.concat([corridors.head(TOP_N), corridors.tail(TOP_N)]).iterrows():
        ws1.append([label, round(row["mean"], 2), int(row["count"])])
    # A Reference just tells openpyxl "read the chart data from this cell range" -
    # data_ref is the numbers to plot, cats_ref is the labels along the axis. This
    # is the same thing Excel does behind the scenes when you select cells and
    # click Insert > Chart.
    chart1 = BarChart()
    chart1.title = "Corridor cost comparison (%)"
    data_ref = Reference(ws1, min_col=2, min_row=1, max_row=ws1.max_row)
    cats_ref = Reference(ws1, min_col=1, min_row=2, max_row=ws1.max_row)
    chart1.add_data(data_ref, titles_from_data=True)
    chart1.set_categories(cats_ref)
    ws1.add_chart(chart1, "E2")

    ws2 = wb.create_sheet("Provider Comparison")
    ws2.append(["Firm type", "Avg total cost %"])
    for cell in ws2[1]:
        cell.font = header_font
    for label, value in provider_avg.items():
        ws2.append([label, round(value, 2)])
    chart2 = BarChart()
    chart2.title = "Average cost by provider type (%)"
    data_ref = Reference(ws2, min_col=2, min_row=1, max_row=ws2.max_row)
    cats_ref = Reference(ws2, min_col=1, min_row=2, max_row=ws2.max_row)
    chart2.add_data(data_ref, titles_from_data=True)
    chart2.set_categories(cats_ref)
    ws2.add_chart(chart2, "D2")

    ws3 = wb.create_sheet("Cost Trend")
    ws3.append(["Period", *trend.columns])
    for cell in ws3[1]:
        cell.font = header_font
    for period, row in trend.iterrows():
        ws3.append([period, *[round(v, 2) if pd.notna(v) else None for v in row]])
    chart3 = LineChart()
    chart3.title = "Cost trend over time by provider type (%)"
    data_ref = Reference(ws3, min_col=2, max_col=1 + len(trend.columns), min_row=1, max_row=ws3.max_row)
    cats_ref = Reference(ws3, min_col=1, min_row=2, max_row=ws3.max_row)
    chart3.add_data(data_ref, titles_from_data=True)
    chart3.set_categories(cats_ref)
    ws3.add_chart(chart3, "H2")

    ws4 = wb.create_sheet("Corridor Heatmap")
    ws4.append(["Sending \\ Receiving", *heatmap.columns])
    for cell in ws4[1]:
        cell.font = header_font
    for source, row in heatmap.iterrows():
        ws4.append([source, *[round(v, 2) if pd.notna(v) else None for v in row]])
    last_col = ws4.cell(row=1, column=ws4.max_column).column_letter
    # ColorScaleRule is the code version of Excel's Home > Conditional Formatting >
    # Colour Scales - green for the lowest value in the range, yellow midway,
    # red for the highest, applied cell-by-cell across the block below.
    rule = ColorScaleRule(
        start_type="min", start_color="63BE7B",
        mid_type="percentile", mid_value=50, mid_color="FFEB84",
        end_type="max", end_color="F8696B",
    )
    ws4.conditional_formatting.add(f"B2:{last_col}{ws4.max_row}", rule)

    wb.save(WORKBOOK_PATH)


def main() -> None:
    df = load_data()
    corridors = chart1_top_corridors(df)
    provider_avg = chart2_provider_comparison(df)
    trend = chart3_cost_trend(df)
    heatmap = chart4_corridor_heatmap(df)
    chart5_cheapest_vs_expensive(df, corridors)
    chart6_choropleth(df)
    chart7_payment_flow_diagram(df)
    build_workbook(corridors, provider_avg, trend, heatmap)

    print("Cheapest corridor:", corridors.index[0], round(corridors.iloc[0]["mean"], 2))
    print("Most expensive corridor:", corridors.index[-1], round(corridors.iloc[-1]["mean"], 2))
    print("Provider averages:\n", provider_avg)


if __name__ == "__main__":
    main()
